import openpyxl
import os
os.chdir('g:')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
wb2=openpyxl.Workbook()
sheet2=wb2.active
sheet2.title = 'final population list'
sheet2.cell(row=1,column=1).value='county'
sheet2.cell(row=1,column=2).value= 'no. of tracts'
sheet2.cell(row=1,column=3).value = 'Population'
countyTracts = 1
j=2
countyPop = sheet.cell(row= 2,column= 4).value
countyName= sheet.cell(row= 2,column= 3).value
for i in range(3,int(sheet.max_row)):
	if sheet.cell(row= i,column=3).value==countyName:
		countyTracts+=1
		countyPop += sheet.cell(row= i,column= 4).value
	else:
		sheet2.cell(row=j,column=1).value= countyName
		sheet2.cell(row=j,column=2).value= countyTracts
		sheet2.cell(row=j,column=3).value= countyPop
		j+=1
		countyName = sheet.cell(row= i,column=3).value
		countyTracts = 1
		countyPop = sheet.cell(row= i,column=4).value
wb2.save('results.xlsx')


	
